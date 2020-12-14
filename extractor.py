import xlrd
import pandas as pd
from openpyxl import load_workbook
import time
import os
import numpy as np
import warnings
import msoffcrypto
import datetime
from datetime import timedelta
import decimal
import mylib
warnings.simplefilter(action='ignore', category=Warning)
settings = mylib.ReadCSV(os.path.join(os.getcwd(), 'settings.csv'))
settings = pd.DataFrame(settings)
sections = []
HRU = []
SR = []
APS = []
dataConfig = [HRU,SR,APS]
fileName = ''
t = ''

def halfRoundUp(data):
    try:
        return int(decimal.Decimal(data*100).quantize(decimal.Decimal('1'), rounding=decimal.ROUND_HALF_UP))
    except:
        return data

def simpleRounding(data):
    try:
        return round(data)
    except:
        return data

def addPercentage(data):
    p = str(data) + '%'
    return p

def dateFormating(data):
    try:
        d = data.strftime('%m/%d/%Y')
        return d
    except:
        return data
    

for i in range(len(settings.columns)):
    # print(settings[i][0])
    if settings[i][0] == 'Sections':
        for a in range(1,len(settings[i])):
            if settings[i][a]:
                sections.append(settings[i][a])
    elif settings[i][0] == '﻿FileName':
        fileName = settings[i][1]
    elif settings[i][0] == 'Output Folder':
        OutputFolder = os.path.join(os.getcwd(), settings[i][1])
    elif settings[i][0] == 'Password':
        password = settings[i][1]
    elif settings[i][0] == 'Time':
        te = settings[i][1]
    elif settings[i][0] == 'Half Round Up':
        for a in range(1,len(settings[i])):
            if settings[i][a]:
                HRU.append(settings[i][a])
    elif settings[i][0] == 'Simple Rounding':
        for a in range(1,len(settings[i])):
            if settings[i][a]:
                SR.append(settings[i][a])
    elif settings[i][0] == 'Add Percentage Sign':
        for a in range(1,len(settings[i])):
            if settings[i][a]:
                APS.append(settings[i][a])
    elif settings[i][0] == 'Header Row':
        hr = int(settings[i][1]) - 1

inputFilePath = '{}/decrypted.xlsx'.format(os.path.dirname(os.path.abspath(__file__)))
while True:
    if not fileName:
        print('No file name found')
        break
    dt = datetime.datetime.utcnow()+timedelta(hours=-8)
    print('Current time is {}.'.format(dt),end="\r", flush=True)
    time.sleep(1)
    if te:
        if dt.hour == int(te):
            t = True
        else:
            t = False
    else:
        t = True
    if t:
        try:
            file = msoffcrypto.OfficeFile(open(fileName, "rb"))
            file.load_key(password=password)
            file.decrypt(open("decrypted.xlsx", "wb"))
            wb = load_workbook(inputFilePath, data_only=True)
            print('\n-- Spreadsheet decrypted --')
        except:
            wb = load_workbook(os.path.join(os.getcwd(), fileName), data_only=True)
            print('\n-- Spreadsheet Loaded --')
        sh = wb.worksheets[wb.worksheets.index(wb['Web Portal Export'])]
        data = pd.DataFrame(sh.values)
        print(data)
        for index, i in enumerate(dataConfig):
            for a in i:

                for i in range(len(data.columns)):
                    for r in range(len(data[i])):
                        data[i][r] = dateFormating(data[i][r])

                for i in range(len(data.columns)):
                    for r in range(len(data[i])):
                        if data[i][r] == None:
                            data[i][r] = ''

                    if a == data[i][1]:
                        for r in range(2,len(data[i])):
                            if data[i][r] != '':
                                if index == 0:
                                    data[i][r] = halfRoundUp(data[i][r])
                                elif index == 1:
                                    data[i][r] = simpleRounding(data[i][r])
                                elif index == 2:
                                    data[i][r] = addPercentage(data[i][r])
                            
        for q in sections:
            df = pd.DataFrame()
            sec = []
            for i in range(len(data.columns)):
                try:
                    if data[i][hr].split('-')[0] == q:
                        for w in data[i]:
                            df[data[i][hr]] = data[i]
                except Exception as e:
                    print(e)
                    pass
            print(df)
            if not os.path.exists(f"{OutputFolder}"):
                try:
                    os.system(f'mkdir {OutputFolder}')
                except Exception as e:
                    print(e)
            outputFilePath = os.path.join(OutputFolder, q + '.xlsx')
            writer = pd.ExcelWriter(outputFilePath, engine='openpyxl')
            df.to_excel(writer,index=False,header=False)
            writer.close()
        try:
            os.system('rm decrypted.xlsx')
        except:
            pass

        print('Export Finished!')
        time.sleep(7200)